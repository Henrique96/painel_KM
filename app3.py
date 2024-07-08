import openpyxl

book = openpyxl.Workbook()

print(book.sheetnames)

book.create_sheet('Frutas', 0)

frutas_page = book['Frutas']
frutas_page.append(['Fruta', 'Quantidade', 'Preço'])
frutas_page.append(['Banana', '5', 'R$3,90'])
frutas_page.append(['Fruta 2', '2', 'R$15,90'])
frutas_page.append(['Fruta 3', '10', 'R$30,90'])
frutas_page.append(['Fruta 4', '2', 'R$50,90'])

book.save('app/Planilha de Compras.xlsx')